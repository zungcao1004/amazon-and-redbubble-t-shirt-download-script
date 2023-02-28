import os
import openpyxl
import tkinter as tk
from tkinter import filedialog
import shutil

# Create a Tkinter root window
root = tk.Tk()
root.withdraw()

# Use the file dialog to ask the user to select a directory
directory = filedialog.askdirectory()

# Create a new workbook object
workbook = openpyxl.Workbook()
sheet = workbook.active

# Open the "All datas redbubble.txt" file for reading
with open('All datas redbubble.txt', 'r', encoding='utf-8') as input_file:
    # Read the lines of the file
    lines = input_file.readlines()
    # Loop through each set of 5 lines (title, url, srcMockup, srcThumbnail, srcDesign)
    for i in range(0, len(lines), 5):
        # Extract the data from the lines
        title = lines[i].strip()
        url = lines[i+1].strip()
        src_mockup = lines[i+2].strip()
        src_thumbnail = lines[i+3].strip()
        src_design = lines[i+4].strip()
        # Write the data to the spreadsheet
        row = i//5 + 1
        sheet.cell(row=row, column=1).value = title
        sheet.cell(row=row, column=2).value = url
        sheet.cell(row=row, column=3).value = src_mockup
        sheet.cell(row=row, column=4).value = src_thumbnail
        sheet.cell(row=row, column=5).value = src_design

# Save the workbook to a file
workbook.save(os.path.join(directory, 'All datas redbubble.xlsx'))

# Move the file to the current directory
filename = os.path.basename('All datas redbubble.xlsx')
shutil.move(os.path.join(directory, filename), filename)
