import os
import openpyxl
import tkinter as tk
from tkinter import filedialog
import shutil

# Create a Tkinter root window
root = tk.Tk()
root.withdraw()

# Use the file dialog to ask the user to select a directory
directory = filedialog.askdirectory()

# Create a new workbook object
workbook = openpyxl.Workbook()
sheet = workbook.active

# Open the "All datas merch.txt" file for reading
with open('All datas merch.txt', 'r') as input_file:
    # Read the lines of the file
    lines = input_file.readlines()
    # Loop through each pair of lines (name and URL)
    for i in range(0, len(lines), 2):
        # Extract the name and URL
        name = lines[i].strip()
        url = lines[i+1].strip()
        # Write the data to the spreadsheet
        sheet.cell(row=i//2+1, column=1).value = name
        sheet.cell(row=i//2+1, column=2).value = url

# Save the workbook to a file
workbook.save(os.path.join(directory, 'All datas merch.xlsx'))

# Move the file to the current directory
filename = os.path.basename('All datas merch.xlsx')
shutil.move(os.path.join(directory, filename), filename)
